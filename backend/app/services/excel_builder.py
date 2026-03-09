"""
services/excel_builder.py — 生成 BOM + 经济分析 Excel 报告
直接从 api.py 的 _build_excel 提取，无 HTTP 相关代码。
"""
from __future__ import annotations
import io
import math
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── 颜色常量 ──────────────────────────────────────────────────
NAVY        = "003C71"
SUBHDR      = "1F4E79"
LIGHT_BLUE  = "DEEAF1"
LIGHT_GREEN = "E2EFDA"
LIGHT_GRAY  = "F2F2F2"
WHITE       = "FFFFFF"
ACCENT      = "FFC000"


def _cfont(size=11, bold=False, color="000000", italic=False) -> Font:
    return Font(name="微软雅黑", size=size, bold=bold, color=color, italic=italic)


def _sfill(hex_c) -> PatternFill:
    return PatternFill("solid", fgColor=hex_c)


def _tborder(style="thin") -> Border:
    s = Side(style=style, color="8EA9C1")
    return Border(left=s, right=s, top=s, bottom=s)


def _cell_set(cell, val, font=None, fill_=None, align=None, border=None):
    cell.value = val
    if font:   cell.font      = font
    if fill_:  cell.fill      = fill_
    if align:  cell.alignment = align
    if border: cell.border    = border


def _set_col_w(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(req) -> bytes:
    """
    生成仿 '解决方案部件清单' 模板格式的 Excel：
      Sheet 1 – BOM（关键零部件清单）
      Sheet 2 – 经济分析（CAPEX 明细 + 20 年年度对比）
      Sheet 3 – 联系信息
    """
    sc  = req.systemConfig   or {}
    cx  = req.capex          or {}
    sim = req.simulation     or {}
    sm  = req.summary        or {}
    ct  = req.comparisonTable or []
    con = req.contact
    now = datetime.datetime.now().strftime("%Y-%m-%d")

    pv_sets       = int(sc.get("bracketSets",       0))
    pv_pps        = int(sc.get("panelsPerSet",      32))
    pv_total      = pv_sets * pv_pps
    pv_kw         = float(sc.get("pvCapacityKw",    0))
    pv_model      = str(sc.get("panelModel",        "—"))
    pv_watts      = int(sc.get("panelWatts",        655))
    bat_count     = int(sc.get("batteryPackCount",   0))
    bat_kwh_total = float(sc.get("batteryCapacityKwh", 0))
    bat_kwh_each  = round(bat_kwh_total / bat_count, 1) if bat_count else 16
    bat_model     = str(sc.get("batteryModel",      "LFP-16kWh"))
    gen_kw        = float(sc.get("dieselCapacityKw", 0))
    gen_model     = str(sc.get("dieselModel",       "—"))
    volt_level    = str(sc.get("voltageLevel",      "48V"))
    ems_mode      = str(sc.get("emsMode",           "edge"))
    ann_load      = float(sc.get("annualLoadKwh",    0))
    area_m2       = float(sc.get("occupiedAreaM2",   0))
    num_inv       = max(1, math.ceil(bat_count / 6))

    if pv_kw > 0 and gen_kw > 0: project_type = "光储柴微电网"
    elif pv_kw > 0:               project_type = "光储微电网"
    else:                         project_type = "离网微电网"

    pv_mod_cost  = float(cx.get("pvModuleCost",        0))
    mount_cost   = float(cx.get("pvMountingCost",       0))
    bess_cost    = float(cx.get("energyStorageCost",    0))
    gen_cost     = float(cx.get("dieselGeneratorCost",  0))
    trans_cost   = float(cx.get("intlTransportCost",    0))
    inst_cost    = float(cx.get("installationCost",     0))
    acc_cost     = float(cx.get("accessoryCost",        0))
    other_cost   = float(cx.get("otherInitialCost",     0))
    subtotal     = float(cx.get("equipmentSubtotal",    0))
    profit_pct   = float(cx.get("profitMargin",         0))
    profit_amt   = float(cx.get("profitAmount",         0))
    selling      = float(cx.get("sellingPrice",         0))

    wb = Workbook()

    # ══ Sheet 1: BOM ═════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "BOM"
    _set_col_w(ws1, [8, 20, 20, 26, 30, 16, 10, 10, 38])

    ws1.row_dimensions[1].height = 60
    ws1.merge_cells("A1:I1")
    _cell_set(ws1["A1"], "关键零部件清单\nKey Components List",
              font=_cfont(20, bold=True, color=NAVY),
              align=Alignment(horizontal="center", vertical="center", wrap_text=True))

    ws1.row_dimensions[2].height = 30
    ws1.merge_cells("A2:G2")
    _cell_set(ws1["A2"],
              f"项目名称 Project Name:   {con.firstName} {con.lastName}  /  {con.company}",
              font=_cfont(11, color="2F5496"),
              align=Alignment(horizontal="left", vertical="center"),
              fill_=_sfill(LIGHT_BLUE))
    ws1.merge_cells("H2:I2")
    _cell_set(ws1["H2"],
              f"产品类别 Product Type: {project_type}",
              font=_cfont(11, color="2F5496"),
              align=Alignment(horizontal="left", vertical="center"),
              fill_=_sfill(LIGHT_BLUE))

    ws1.row_dimensions[3].height = 52
    hdr_cols = [
        "序号\nNO.", "零件编码\nPart Number",
        "零件中文名称\nPart Chinese Name", "零件英文名称\nPart English Name",
        "规格型号\nSpecification and Models", "尺寸(mm)\nDimensions",
        "数量\nQuantity", "单位\nUnit", "规格/型号\nSpecification / Model",
    ]
    for ci, h in enumerate(hdr_cols, 1):
        cell = ws1.cell(3, ci, h)
        cell.font      = _cfont(11, bold=True, color="FFFFFF")
        cell.fill      = _sfill("1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _tborder()

    def _dim_pv(w: int) -> str:
        return "2384×1303×33mm" if w >= 655 else "—"

    bom_data = [
        (1, "", "混合逆变器", "Hybrid Inverter", "18K-2P-LV", "",
         num_inv, "EA",
         f"Input: PV + Battery ({volt_level}), Output: AC 230V/50Hz; supports off-grid & grid-tied modes"),
        (2, "", "电池包", "Battery Pack", bat_model, "",
         bat_count, "EA",
         f"LFP, {bat_kwh_each} kWh/pack, total {bat_kwh_total:.0f} kWh; cycle life ≥ 4,000 cycles; BMS included"),
        (3, "", "工控机 / EMS", "Industrial PC / EMS", "ECU-1170", "",
         1, "EA",
         f"VoltageEnergy EMS; control mode: {ems_mode}; real-time monitoring & remote O&M"),
        (4, "ES1-S002-1", "一体化托盘", "Integrated Skid", "", "",
         pv_sets, "SET",
         "Main material: Q355B hot-dip galvanized; integrated pallet spliced weldment"),
        (5, "ES1-S001-1", "折叠支架系统", "Foldable Mounting System", "", "",
         pv_sets, "SET",
         f"A set consisting of {pv_pps} rows of PV modules; steel-aluminum structure; foldable design"),
        (6, "", "BOS箱", "BOS Box", "", "",
         pv_sets, "EA",
         "Balance of System box; includes DC/AC breakers, combiner, surge protection, metering"),
        (7, "ES1-E004-1", "光伏组件", "PV Panel Module", pv_model, _dim_pv(pv_watts),
         pv_total, "EA",
         f"{pv_model}; {pv_watts}Wp; total {pv_kw:.1f} kW; footprint ≈ {area_m2:.0f} m²"),
    ]
    if gen_kw > 0:
        bom_data.append((
            8, "ES1-E005-1", "柴油发电机", "Diesel Generator",
            gen_model if gen_model and gen_model != "—" else f"{gen_kw:.0f}kW genset", "",
            1, "EA",
            f"{gen_kw:.0f} kW; backup power for cloudy days; annual run hours reduced "
            f"~{(1 - float(sim.get('mgDieselHours', 0)) / max(float(sim.get('dieselRunHoursA', 8760)), 1)) * 100:.0f}% vs diesel-only",
        ))

    for idx, row_data in enumerate(bom_data):
        ri = 4 + idx
        no_, pn, zh, en, spec, dims, qty, unit, notes = row_data
        bg = _sfill(LIGHT_GRAY if idx % 2 == 0 else WHITE)
        for ci, val in enumerate([no_, pn, zh, en, spec, dims, qty, unit, notes], 1):
            cell = ws1.cell(ri, ci, val)
            cell.fill      = bg
            cell.border    = _tborder()
            cell.font      = _cfont(11, bold=(ci in (3, 4)))
            cell.alignment = Alignment(
                horizontal="center" if ci in (1, 6, 7, 8) else "left",
                vertical="center", wrap_text=True)
        ws1.row_dimensions[ri].height = 42

    sig_row = 4 + len(bom_data) + 1
    ws1.row_dimensions[sig_row].height = 30
    ws1.merge_cells(f"A{sig_row}:I{sig_row}")
    _cell_set(ws1[f"A{sig_row}"],
              f"制单 Prepared by:              审核 Checked by:              "
              f"批准 Approved by:              日期 Date: {now}",
              font=_cfont(10, color="595959"),
              align=Alignment(horizontal="left", vertical="center"),
              fill_=_sfill(LIGHT_BLUE))

    note_row = sig_row + 1
    ws1.merge_cells(f"A{note_row}:I{note_row}")
    ws1.row_dimensions[note_row].height = 20
    _cell_set(ws1[f"A{note_row}"],
              "备注 Notes: All specifications subject to final purchase order confirmation. "
              "Prices in USD. Quantities to be verified on-site.",
              font=_cfont(9, italic=True, color="808080"),
              align=Alignment(horizontal="left", vertical="center"))

    # ══ Sheet 2: 经济分析 ════════════════════════════════════════
    ws2 = wb.create_sheet("经济分析")
    _set_col_w(ws2, [32, 22, 20, 20, 20, 20, 20, 20, 20])

    ws2.merge_cells("A1:B1")
    ws2.row_dimensions[1].height = 26
    _cell_set(ws2["A1"], "CAPEX 成本明细  /  Capital Cost Breakdown",
              font=_cfont(13, bold=True, color="FFFFFF"), fill_=_sfill(SUBHDR),
              align=Alignment(horizontal="center", vertical="center"))

    capex_rows = [
        ("光伏组件  PV Modules",            pv_mod_cost),
        ("折叠支架  Mounting System",        mount_cost),
        ("储能系统  Battery Storage (BESS)", bess_cost),
        ("柴油发电机  Diesel Generator",     gen_cost),
        ("国际运输  International Freight",  trans_cost),
        ("安装费  Installation",             inst_cost),
        ("附件材料  Accessories",            acc_cost),
        ("其他初始费用  Other Initial",      other_cost),
    ]
    for ri, (label, amt) in enumerate(capex_rows, 2):
        ws2.row_dimensions[ri].height = 18
        bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
        for ci, val in enumerate([label, f"$ {amt:,.0f}" if amt else "—"], 1):
            cell = ws2.cell(ri, ci, val)
            cell.font   = _cfont(11, bold=(ci == 1))
            cell.fill   = _sfill(bg)
            cell.border = _tborder()
            cell.alignment = Alignment(horizontal="right" if ci == 2 else "left", vertical="center")

    sub_row = len(capex_rows) + 2
    for ri, (label, val, bg_c) in enumerate([
        ("设备小计  Equipment Subtotal",              f"$ {subtotal:,.0f}",  LIGHT_BLUE),
        (f"利润  Profit Margin ({profit_pct:.1f}%)",  f"$ {profit_amt:,.0f}", WHITE),
        ("含利润报价  Quote (incl. Margin)",           f"$ {selling:,.0f}",   ACCENT),
    ], sub_row):
        ws2.row_dimensions[ri].height = 20
        for ci, txt in enumerate([label, val], 1):
            cell = ws2.cell(ri, ci, txt)
            cell.font   = _cfont(11, bold=True, color=NAVY if "报价" in label else "000000")
            cell.fill   = _sfill(bg_c)
            cell.border = _tborder()
            cell.alignment = Alignment(horizontal="right" if ci == 2 else "left", vertical="center")

    es_row = sub_row + 5
    ws2.merge_cells(f"A{es_row}:B{es_row}")
    ws2.row_dimensions[es_row].height = 24
    _cell_set(ws2[f"A{es_row}"], "经济效益摘要  /  Economic Summary",
              font=_cfont(13, bold=True, color="FFFFFF"), fill_=_sfill(SUBHDR),
              align=Alignment(horizontal="center", vertical="center"))

    econ_items = [
        ("年用电量  Annual Load",           f"{ann_load:,.0f} kWh"),
        ("太阳能保障率  Solar Fraction",     f"{sim.get('solarFractionPct', 0):.1f}%"),
        ("年节省燃油量  Fuel Savings",       f"{sim.get('annualFuelSavingLiters', 0):,.0f} L/yr"),
        ("年节省费用  Annual Cost Savings",  f"$ {sim.get('annualFuelSavingUsd', 0):,.0f}"),
        ("微电网年运维  MG Annual O&M",      f"$ {sm.get('mgAnnualOmUsd', 0):,.0f}"),
        ("微电网 LCOE",                      f"$ {sm.get('finalMgLcoe', 0):.4f}/kWh"),
        ("柴油 LCOE  Diesel-Only LCOE",     f"$ {sm.get('finalDieselLcoe', 0):.4f}/kWh"),
        ("投资回本年限  Payback Period",     f"Year {sm.get('breakevenYear', '—')}"),
        ("20 年累计收益  20-yr Net Revenue", f"$ {sm.get('finalCumulativeRevenue', 0):,.0f}"),
    ]
    for ri, (k, v) in enumerate(econ_items, es_row + 1):
        bg = LIGHT_GREEN if k.startswith("投资") or k.startswith("20") else (LIGHT_GRAY if ri % 2 == 0 else WHITE)
        ws2.row_dimensions[ri].height = 18
        for ci, val in enumerate([k, v], 1):
            cell = ws2.cell(ri, ci, val)
            cell.font   = _cfont(11, bold=(k.startswith("投资") or k.startswith("20")))
            cell.fill   = _sfill(bg)
            cell.border = _tborder()
            cell.alignment = Alignment(horizontal="right" if ci == 2 else "left", vertical="center")

    if ct:
        yt_row = es_row + len(econ_items) + 3
        ws2.merge_cells(f"A{yt_row}:I{yt_row}")
        ws2.row_dimensions[yt_row].height = 24
        _cell_set(ws2[f"A{yt_row}"],
                  "微电网 vs 纯柴油 年度成本对比 / Annual Cost Comparison (MG vs Diesel-Only)",
                  font=_cfont(12, bold=True, color="FFFFFF"), fill_=_sfill(SUBHDR),
                  align=Alignment(horizontal="center", vertical="center"))
        y_hdrs = ["年份\nYear", "微电网年费用\nMG Annual", "柴油年费用\nDiesel Annual",
                  "微电网累计\nMG Cumul.", "柴油累计\nDiesel Cumul.",
                  "MG LCOE", "柴油 LCOE", "年度收益\nAnnual Rev.", "累计收益\nCumul. Rev."]
        for ci, h in enumerate(y_hdrs, 1):
            cell = ws2.cell(yt_row + 1, ci, h)
            cell.font      = _cfont(10, bold=True, color="FFFFFF")
            cell.fill      = _sfill("2E74B5")
            cell.border    = _tborder()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws2.row_dimensions[yt_row + 1].height = 36

        bk_yr = sm.get("breakevenYear")
        for di, row in enumerate(ct):
            ri = yt_row + 2 + di
            is_bk = bk_yr is not None and row.get("year") == bk_yr
            bg = _sfill(LIGHT_GREEN) if is_bk else _sfill(LIGHT_GRAY if di % 2 == 0 else WHITE)
            vals = [
                row.get("year"),
                f"${row.get('mgAnnualCost', 0):,.0f}",
                f"${row.get('dieselAnnualCost', 0):,.0f}",
                f"${row.get('mgCumulative', 0):,.0f}",
                f"${row.get('dieselCumulative', 0):,.0f}",
                f"${row.get('mgLcoe', 0):.4f}",
                f"${row.get('dieselLcoe', 0):.4f}",
                f"${row.get('annualRevenue', 0):,.0f}",
                f"${row.get('cumulativeRevenue', 0):,.0f}",
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws2.cell(ri, ci, v)
                cell.font   = _cfont(10, bold=is_bk)
                cell.fill   = bg
                cell.border = _tborder()
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws2.row_dimensions[ri].height = 16

    # ══ Sheet 3: 联系信息 ════════════════════════════════════════
    ws3 = wb.create_sheet("联系信息")
    _set_col_w(ws3, [28, 36])
    ws3.merge_cells("A1:B1")
    ws3.row_dimensions[1].height = 26
    _cell_set(ws3["A1"], "客户联系信息  /  Customer Contact Information",
              font=_cfont(13, bold=True, color="FFFFFF"), fill_=_sfill(SUBHDR),
              align=Alignment(horizontal="center", vertical="center"))

    ctc_rows = [
        ("姓名  Full Name",        f"{con.firstName} {con.lastName}"),
        ("公司  Company",          con.company),
        ("电子邮箱  Email",        con.email),
        ("联系电话  Phone",        con.phone),
        ("州 / 地区  State",       con.state),
        ("城市  City",             con.city),
        ("报告日期  Report Date",  now),
        ("生成系统  System",       "VoltageEnergy™ Microgrid Advisor v2.0"),
    ]
    for ri, (k, v) in enumerate(ctc_rows, 2):
        bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
        ws3.row_dimensions[ri].height = 20
        for ci, val in enumerate([k, v], 1):
            cell = ws3.cell(ri, ci, val)
            cell.font   = _cfont(11, bold=(ci == 1))
            cell.fill   = _sfill(bg)
            cell.border = _tborder()
            cell.alignment = Alignment(vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
